from django.shortcuts import render
from django.http import FileResponse  # Добавлено
import os
from django.conf import settings
from openpyxl import load_workbook
from docx import Document
from django.core.files.storage import FileSystemStorage

def fill_word_template(excel_file_path, word_template_path, output_path):
    try:
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active
        data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data[row[0]] = str(row[2])
        doc = Document(word_template_path)
        for paragraph in doc.paragraphs:
            for key, value in data.items():
                if '{{' + key + '}}' in paragraph.text:
                    paragraph.text = paragraph.text.replace('{{' + key + '}}', value)
        doc.save(output_path)
        return True
    except Exception as e:
        print(f"Error processing files: {e}")
        return False

def upload_file(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if excel_file:
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            excel_file_path = os.path.join(settings.MEDIA_ROOT, filename)
            output_filename = filename.rsplit('.', 1)[0] + '.docx'
            word_filepath = os.path.join(settings.MEDIA_ROOT, output_filename)

            if fill_word_template(excel_file_path, os.path.join(settings.BASE_DIR, 'spravka/templates/template.docx'), word_filepath):
                os.remove(excel_file_path)

                return render(request, 'document_processing/upload_success.html', {
                    'download_link': output_filename,
                })
            else:
                return render(request, 'document_processing/upload_form.html', {'error': 'Ошибка при обработке файла'})
        else:
            return render(request, 'document_processing/upload_form.html', {'error': 'Файл не выбран'})

    return render(request, 'document_processing/upload_form.html')

def download_template(request):
    file_path = os.path.join(settings.BASE_DIR, 'document_processing/static/document_processing/template.xlsx')
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='template.xlsx')

def download_file(request, filename):
    file_path = os.path.join(settings.MEDIA_ROOT, filename)
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=filename)